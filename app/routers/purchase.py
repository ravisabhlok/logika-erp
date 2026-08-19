import io
import re
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import RedirectResponse, StreamingResponse, Response
from fastapi.templating import Jinja2Templates
from markupsafe import Markup, escape
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload
from xhtml2pdf import pisa

from app.database import get_db
from app.auth import require_module_permission, get_user_module_permissions, require_admin
from app.formatting import format_qty, format_inr, format_money, format_dt, IST_OFFSET, CURRENCY_SYMBOLS
from app.models import (
    PurchaseOrder, PurchaseOrderItem, Vendor, Item, StockTransaction, ItemSerial, User, Company,
)
from app.requirements import compute_demand_map, on_order_qty, last_vendor
from app.audit import diff_fields, log_field_changes, log_action

router = APIRouter(prefix="/purchase", tags=["purchase"])
templates = Jinja2Templates(directory="app/templates")
templates.env.filters["qty"] = format_qty
templates.env.filters["inr"] = format_inr
templates.env.filters["money"] = format_money
templates.env.filters["dt"] = format_dt


def _nl2br(value):
    """Same xhtml2pdf quirk/fix as sales.py's _nl2br — see that function's
    docstring (each router keeps its own small Jinja2Templates instance and
    filter set in this app rather than a shared central environment, so
    this is deliberately duplicated here, not imported)."""
    if not value:
        return ""
    lines = [str(escape(line)) for line in str(value).split("\n")]
    return Markup("<br/>".join(lines))


templates.env.filters["nl2br"] = _nl2br


def next_order_no(db: Session) -> str:
    """PO-00001, PO-00002, ... — derived from the highest existing numeric
    suffix, not a row count. A row count breaks the moment any order other
    than the most recent one is deleted: the count drops but the highest
    number in use doesn't, so the next generated number can collide with
    one that's still there (this is what caused the duplicate-key error on
    2026-08-19 — a deleted draft order left a gap)."""
    existing = db.query(PurchaseOrder.order_no).filter(PurchaseOrder.order_no.like("PO-%")).all()
    max_n = 0
    for (order_no,) in existing:
        suffix = order_no[len("PO-"):]
        if suffix.isdigit():
            max_n = max(max_n, int(suffix))
    return f"PO-{max_n + 1:05d}"


def compute_global_requirements(db: Session):
    """Company-wide "what do I need to order" — the one number this app
    gives you, rather than one that only looks at a single sales order in
    isolation (which would double-book stock that another open order is
    also counting on).

    Demand is aggregated from every open ('draft' or 'confirmed') sales
    order at once — not just one — so an item shared across several orders
    is netted against stock a single time. Every item carrying a
    `reorder_level` is also seeded into the same computation with its
    buffer added on top of whatever sales demand it has (zero, if none) —
    so plain minimum-stock replenishment shows up here too, not just
    order-driven shortages. Assemblies are exploded down through their BOM
    the same way as the per-order view used to. See app/requirements.py for
    the shared walk.

    Does not reserve stock for 'delivered' orders (already deducted) or
    'cancelled' ones (never will be).
    """
    demand, sources, exploded = compute_demand_map(db)

    results = []
    for item_id, sales_demand in demand.items():
        if item_id in exploded:
            continue  # assemblies were broken down into their components above
        item = db.query(Item).get(item_id)
        if not item:
            continue
        buffer = float(item.reorder_level or 0)
        required = sales_demand + buffer
        in_stock = float(item.current_stock or 0)
        on_order = on_order_qty(db, item_id)
        shortfall = max(0.0, required - in_stock - on_order)
        if shortfall > 0:
            results.append({
                "item": item,
                "sales_demand": sales_demand,
                "reorder_level": buffer,
                "in_stock": in_stock,
                "on_order": on_order,
                "shortfall": shortfall,
                "orders": sorted(sources.get(item_id, set())),
                "last_vendor": last_vendor(db, item_id),
            })

    return sorted(results, key=lambda r: r["item"].name)


@router.get("")
def list_purchase(request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "view"))):
    orders = db.query(PurchaseOrder).options(joinedload(PurchaseOrder.vendor)).order_by(PurchaseOrder.created_at.desc()).all()
    perms = get_user_module_permissions(user, db, "purchase")
    return templates.TemplateResponse("purchase/list.html", {"request": request, "user": user, "orders": orders, "perms": perms})


@router.get("/requirements")
def purchase_requirements(request: Request, highlight: str = "", db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "view"))):
    requirements = compute_global_requirements(db)
    perms = get_user_module_permissions(user, db, "purchase")
    return templates.TemplateResponse(
        "purchase/requirements.html",
        {"request": request, "user": user, "requirements": requirements, "highlight": highlight, "perms": perms},
    )


@router.get("/requirements/export")
def export_requirements(db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "view"))):
    """Same "What to Order" computation as the page above, as a downloadable
    .xlsx — same permission (view) since it's just a different format of
    data the user can already see on screen, not a new capability. Numbers
    are written as real numeric cells (not the |qty-formatted strings the
    HTML table shows) so the sheet can be summed/filtered/sorted directly in
    Excel rather than needing to be re-parsed from text."""
    requirements = compute_global_requirements(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "What to Order"

    headers = [
        "Item", "Sales Demand", "Min Stock", "In Stock",
        "On Order", "Still to Order", "Driven By", "Last Vendor",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for r in requirements:
        ws.append([
            r["item"].name,
            float(r["sales_demand"]),
            float(r["reorder_level"]),
            float(r["in_stock"]),
            float(r["on_order"]),
            float(r["shortfall"]),
            ", ".join(r["orders"]) if r["orders"] else "Min stock",
            r["last_vendor"].name if r["last_vendor"] else "",
        ])

    column_widths = [38, 14, 12, 12, 12, 14, 26, 24]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Same date+time-stamped naming convention as scripts/backup_db.py, in
    # IST like every other timestamp shown in this app — one export per
    # click, never silently overwriting a previous download.
    stamp = (datetime.utcnow() + IST_OFFSET).strftime("%Y%m%d_%H%M%S")
    filename = f"what_to_order_{stamp}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/new")
def new_purchase_form(request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "add"))):
    vendors = db.query(Vendor).order_by(Vendor.name).all()
    items = db.query(Item).filter(Item.is_active == True).order_by(Item.name).all()  # noqa: E712
    return templates.TemplateResponse(
        "purchase/form.html",
        {
            "request": request, "user": user, "vendors": vendors, "items": items,
            "currencies": sorted(CURRENCY_SYMBOLS), "order": None, "submitted": {},
        },
    )


@router.post("/new")
async def create_purchase_order(request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "add"))):
    form = await request.form()
    vendor_id = int(form.get("vendor_id"))
    notes = form.get("notes", "")
    item_ids = form.getlist("item_id")
    quantities = form.getlist("quantity")
    prices = form.getlist("unit_price")
    currency = (form.get("currency") or "INR").strip().upper()[:3] or "INR"
    try:
        exchange_rate = float(form.get("exchange_rate") or 1)
    except ValueError:
        exchange_rate = 1
    if exchange_rate <= 0:
        exchange_rate = 1

    order = PurchaseOrder(
        order_no=next_order_no(db),
        vendor_id=vendor_id,
        notes=notes,
        created_by=user.id,
        status="draft",
        currency=currency,
        exchange_rate=exchange_rate,
    )
    db.add(order)

    total = 0
    for item_id_str, qty_str, price_str in zip(item_ids, quantities, prices):
        if not item_id_str or not qty_str:
            continue
        qty = float(qty_str)
        if qty <= 0:
            continue
        price = float(price_str) if price_str else 0
        item = db.query(Item).get(int(item_id_str))
        line_total = price * qty
        total += line_total
        order.items.append(PurchaseOrderItem(
            item_id=item.id, quantity=qty, unit_price=price, total=line_total,
        ))

    order.total_amount = total
    db.flush()  # need order.id before logging
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "create",
        summary=f"Created for {order.vendor.name if order.vendor else order.vendor_id} — {len(order.items)} line(s), total {order.currency} {order.total_amount}",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order.id}", status_code=303)


@router.get("/{order_id}")
def view_purchase_order(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "view"))):
    order = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.vendor), joinedload(PurchaseOrder.items).joinedload(PurchaseOrderItem.item))
        .get(order_id)
    )
    perms = get_user_module_permissions(user, db, "purchase")
    return templates.TemplateResponse("purchase/detail.html", {"request": request, "user": user, "order": order, "perms": perms})


# Purchase orders this document can be downloaded for. Gated on status, not
# just permission, so the PDF can't be pulled (e.g. by hitting the URL
# directly) before an admin has approved it AND a second person has
# confirmed it — see approve_purchase_order and confirm_purchase_order.
DOWNLOADABLE_PURCHASE_STATUSES = ("ordered", "received")


@router.get("/{order_id}/pdf")
def purchase_order_pdf(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "view"))):
    """Same approach as sales.py's sales_order_pdf: renders purchase/pdf.html
    (a standalone print layout, not base.html) to HTML and converts it to a
    PDF via xhtml2pdf. Only available once the order has been approved
    (status 'ordered' or 'received') — before that there's nothing
    authorized to hand to a vendor yet."""
    order = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.vendor), joinedload(PurchaseOrder.items).joinedload(PurchaseOrderItem.item))
        .get(order_id)
    )
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status not in DOWNLOADABLE_PURCHASE_STATUSES:
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=This purchase order must be approved before it can be downloaded",
            status_code=303,
        )

    company = db.query(Company).first()
    # Same palette as the status badge on purchase/detail.html and
    # purchase/list.html, just as hex since the PDF has no Bootstrap.
    status_color = {
        "draft": "#6c757d", "pending_approval": "#a06a00", "pending_confirmation": "#a06a00",
        "ordered": "#0d6efd", "received": "#198754", "cancelled": "#dc3545",
    }.get(order.status, "#6c757d")
    logo_file = Path(__file__).resolve().parent.parent / "static" / "Logo.jpg"
    logo_path = str(logo_file) if logo_file.exists() else None
    html = templates.get_template("purchase/pdf.html").render(
        order=order, company=company, generated_at=datetime.utcnow(),
        status_color=status_color, logo_path=logo_path, request=request, user=user,
    )
    buffer = io.BytesIO()
    result = pisa.CreatePDF(io.StringIO(html), dest=buffer)
    if result.err:
        return Response(content="Failed to generate PDF", status_code=500)
    buffer.seek(0)
    return Response(
        content=buffer.read(),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'inline; filename="{order.order_no}.pdf"',
            # Regenerated fresh from the DB on every request — see
            # sales_order_pdf's identical comment for why this matters.
            "Cache-Control": "no-store, no-cache, must-revalidate",
        },
    )


# Statuses a purchase order can still be edited from. Stock (and, for
# serialized items, ItemSerial rows) only get written once an order is
# marked 'received' (see receive_purchase_order) — so editing anything up
# to that point can never desync the stock ledger / serial records from
# what's actually on the order. 'received' is where that stops being true:
# line items would no longer match what's already been posted. 'cancelled'
# is allowed too since it never touched stock either; editing one revives
# it back to 'draft' (see update_purchase_order below) rather than leaving
# it edited but still marked cancelled — same convention Sales Orders use,
# see app/routers/sales.py's edit_sales_form. 'pending_approval' and
# 'pending_confirmation' are deliberately NOT in this list — an order
# awaiting approval or confirmation shouldn't be editable out from under
# whoever is reviewing it; reject it back to draft first (see
# reject_purchase_order / reject_purchase_order_confirmation) if it needs
# changes.
EDITABLE_PURCHASE_STATUSES = ("draft", "ordered", "cancelled")


@router.get("/{order_id}/edit")
def edit_purchase_form(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "edit"))):
    order = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.items).joinedload(PurchaseOrderItem.item))
        .get(order_id)
    )
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status not in EDITABLE_PURCHASE_STATUSES:
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Received orders can't be edited — line items would no longer match the stock/serials already recorded",
            status_code=303,
        )

    vendors = db.query(Vendor).order_by(Vendor.name).all()
    # Include any item already on this order even if it's since been marked
    # inactive, so its line doesn't silently disappear from the picker.
    order_item_ids = [line.item_id for line in order.items]
    items = (
        db.query(Item)
        .filter(or_(Item.is_active == True, Item.id.in_(order_item_ids)))  # noqa: E712
        .order_by(Item.name)
        .all()
    )
    submitted = {
        "vendor_id": order.vendor_id,
        "notes": order.notes or "",
        "currency": order.currency,
        "exchange_rate": float(order.exchange_rate),
        "lines": [
            {"item_id": line.item_id, "quantity": float(line.quantity), "unit_price": float(line.unit_price)}
            for line in order.items
        ],
    }
    return templates.TemplateResponse(
        "purchase/form.html",
        {
            "request": request, "user": user, "vendors": vendors, "items": items,
            "currencies": sorted(CURRENCY_SYMBOLS), "order": order, "submitted": submitted,
        },
    )


@router.post("/{order_id}/edit")
async def update_purchase_order(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "edit"))):
    order = db.query(PurchaseOrder).options(joinedload(PurchaseOrder.items)).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status not in EDITABLE_PURCHASE_STATUSES:
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Received orders can't be edited — line items would no longer match the stock/serials already recorded",
            status_code=303,
        )
    was_cancelled = order.status == "cancelled"

    form = await request.form()
    vendor_id = int(form.get("vendor_id"))
    notes = form.get("notes", "")
    item_ids = form.getlist("item_id")
    quantities = form.getlist("quantity")
    prices = form.getlist("unit_price")
    currency = (form.get("currency") or "INR").strip().upper()[:3] or "INR"
    try:
        exchange_rate = float(form.get("exchange_rate") or 1)
    except ValueError:
        exchange_rate = 1
    if exchange_rate <= 0:
        exchange_rate = 1

    changes = diff_fields(order, {
        "vendor_id": vendor_id,
        "notes": notes,
        "currency": currency,
        "exchange_rate": exchange_rate,
    })
    old_line_count = len(order.items)
    old_total = order.total_amount

    order.vendor_id = vendor_id
    order.notes = notes
    order.currency = currency
    order.exchange_rate = exchange_rate
    order.items = []  # cascade="all, delete-orphan" removes the old lines on flush

    total = 0
    for item_id_str, qty_str, price_str in zip(item_ids, quantities, prices):
        if not item_id_str or not qty_str:
            continue
        qty = float(qty_str)
        if qty <= 0:
            continue
        price = float(price_str) if price_str else 0
        item = db.query(Item).get(int(item_id_str))
        line_total = price * qty
        total += line_total
        order.items.append(PurchaseOrderItem(
            item_id=item.id, quantity=qty, unit_price=price, total=line_total,
        ))

    order.total_amount = total
    if len(order.items) != old_line_count or order.total_amount != old_total:
        changes["line_items"] = (
            f"{old_line_count} line(s), total {old_total}",
            f"{len(order.items)} line(s), total {order.total_amount}",
        )
    if was_cancelled:
        # Editing a cancelled order means it's being revived — send it back
        # through the normal lifecycle rather than leaving it edited but
        # still marked cancelled.
        changes["status"] = ("cancelled", "draft")
        order.status = "draft"
    log_field_changes(db, user, "purchase_order", order.id, order.order_no, changes)
    db.commit()
    success_msg = "Purchase order updated" + (" and moved back to draft" if was_cancelled else "")
    return RedirectResponse(url=f"/purchase/{order_id}?success={success_msg}", status_code=303)


@router.post("/{order_id}/status")
def update_purchase_status(
    order_id: int,
    new_status: str = Form(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_module_permission("purchase", "edit")),
):
    """Generic status-change endpoint — deliberately narrow. 'ordered' can
    only be reached through confirm_purchase_order now (after
    approve_purchase_order moves it to 'pending_confirmation'), and
    'received' always needed its own dedicated flow to capture serial
    numbers — so this route only ever actually handles 'cancelled'. Kept as
    its own endpoint (rather than folded away) since Cancel remains a plain,
    unconditional action from draft/pending_approval/pending_confirmation/
    ordered alike, with nothing else to capture."""
    if new_status == "received":
        return RedirectResponse(url=f"/purchase/{order_id}/receive", status_code=303)
    if new_status != "cancelled":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Use Send for Approval / Approve / Reject to change this order's status",
            status_code=303,
        )

    order = db.query(PurchaseOrder).get(order_id)
    old_status = order.status
    order.status = new_status
    if new_status != old_status:
        log_action(db, user, "purchase_order", order.id, order.order_no, "status_change", summary=f"{old_status} -> {new_status}")
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Status updated", status_code=303)


@router.post("/{order_id}/send-for-approval")
def send_purchase_order_for_approval(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_module_permission("purchase", "edit")),
):
    """Staff-facing half of the approval workflow: draft -> pending_approval.
    Same permission as everything else that used to move an order straight
    to 'ordered' — the difference now is it stops one step short and waits
    for an admin."""
    order = db.query(PurchaseOrder).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status != "draft":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only a draft order can be sent for approval",
            status_code=303,
        )
    order.status = "pending_approval"
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "status_change",
        summary=f"Sent for approval by {user.username} — draft -> pending_approval",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Sent for approval", status_code=303)


@router.post("/{order_id}/approve")
def approve_purchase_order(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Admin-facing half: pending_approval -> pending_confirmation.
    Deliberately gated on require_admin (the coarse admin/staff account
    type) rather than the per-module permission matrix — approval is meant
    to be a genuine second set of eyes, not something the same
    Purchase-module-edit permission that got the order to
    'pending_approval' can also grant itself. Approval alone no longer
    unlocks the PDF — a second, independently-permissioned person still has
    to confirm it (see confirm_purchase_order) before it reaches 'ordered'
    and becomes downloadable (see DOWNLOADABLE_PURCHASE_STATUSES)."""
    order = db.query(PurchaseOrder).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status != "pending_approval":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only an order pending approval can be approved",
            status_code=303,
        )
    order.status = "pending_confirmation"
    order.approved_by = user.id
    order.approved_at = datetime.utcnow()
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "approve",
        summary=f"Approved by {user.username} — pending_approval -> pending_confirmation",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Purchase order approved — awaiting confirmation", status_code=303)


@router.post("/{order_id}/reject")
def reject_purchase_order(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """The other admin-facing option: pending_approval -> draft, so an
    order that isn't right yet can go back for changes instead of being
    stuck waiting or getting approved anyway. Same require_admin gate as
    approve_purchase_order."""
    order = db.query(PurchaseOrder).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status != "pending_approval":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only an order pending approval can be rejected",
            status_code=303,
        )
    order.status = "draft"
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "status_change",
        summary=f"Rejected by {user.username} — pending_approval -> draft",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Sent back to draft for changes", status_code=303)


@router.post("/{order_id}/confirm")
def confirm_purchase_order(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_module_permission("purchase", "confirm")),
):
    """Second sign-off, after an admin's approval: pending_confirmation ->
    ordered. Gated on the "confirm" action of the per-module permission
    matrix (RolePermission.can_confirm — see app.auth.ACTIONS), which an
    admin assigns per-Role on the Roles page, independent of the
    admin-only approve/reject gate above. 'ordered' is what unlocks the PDF
    download (see DOWNLOADABLE_PURCHASE_STATUSES)."""
    order = db.query(PurchaseOrder).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status != "pending_confirmation":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only an order pending confirmation can be confirmed",
            status_code=303,
        )
    order.status = "ordered"
    order.confirmed_by = user.id
    order.confirmed_at = datetime.utcnow()
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "confirm",
        summary=f"Confirmed by {user.username} — pending_confirmation -> ordered",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Purchase order confirmed", status_code=303)


@router.post("/{order_id}/reject-confirmation")
def reject_purchase_order_confirmation(
    order_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_module_permission("purchase", "confirm")),
):
    """The other confirm-facing option: pending_confirmation -> draft, so an
    order that isn't right yet can go back for changes (and will need to be
    sent for approval again) instead of being stuck waiting or getting
    confirmed anyway. Same "confirm" permission gate as confirm_purchase_order."""
    order = db.query(PurchaseOrder).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)
    if order.status != "pending_confirmation":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only an order pending confirmation can be rejected",
            status_code=303,
        )
    order.status = "draft"
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "status_change",
        summary=f"Confirmation rejected by {user.username} — pending_confirmation -> draft",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Sent back to draft for changes", status_code=303)


def _split_serials(raw: str) -> list[str]:
    """Split a textarea of serial numbers on commas and/or newlines."""
    parts = re.split(r"[\n,]+", raw or "")
    return [p.strip() for p in parts if p.strip()]


@router.get("/{order_id}/receive")
def receive_form(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "edit"))):
    order = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.vendor), joinedload(PurchaseOrder.items).joinedload(PurchaseOrderItem.item))
        .get(order_id)
    )
    if order.status != "ordered":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only orders in 'ordered' status can be received",
            status_code=303,
        )
    return templates.TemplateResponse(
        "purchase/receive.html",
        {"request": request, "user": user, "order": order, "errors": {}, "submitted": {}},
    )


@router.post("/{order_id}/receive")
async def receive_purchase_order(order_id: int, request: Request, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "edit"))):
    order = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.vendor), joinedload(PurchaseOrder.items).joinedload(PurchaseOrderItem.item))
        .get(order_id)
    )
    if order.status != "ordered":
        return RedirectResponse(
            url=f"/purchase/{order_id}?error=Only orders in 'ordered' status can be received",
            status_code=303,
        )

    form = await request.form()

    # Validate serial counts for every serialized line before writing anything.
    errors = {}
    submitted = {}
    serials_by_line = {}
    for line in order.items:
        field_name = f"serials_{line.id}"
        raw = form.get(field_name, "")
        submitted[line.id] = raw
        if line.item.has_serial:
            serials = _split_serials(raw)
            serials_by_line[line.id] = serials
            if len(serials) != line.quantity:
                errors[line.id] = f"Expected {line.quantity} serial number(s), got {len(serials)}."
            elif len(set(serials)) != len(serials):
                errors[line.id] = "Duplicate serial numbers entered for this line."

    if errors:
        return templates.TemplateResponse(
            "purchase/receive.html",
            {"request": request, "user": user, "order": order, "errors": errors, "submitted": submitted},
            status_code=400,
        )

    # Also reject serials that already exist for that item (e.g. re-entered by mistake).
    for line in order.items:
        if not line.item.has_serial:
            continue
        existing = {
            s for (s,) in db.query(ItemSerial.serial_number).filter(ItemSerial.item_id == line.item_id).all()
        }
        clashes = existing.intersection(serials_by_line[line.id])
        if clashes:
            errors[line.id] = f"Already recorded for this item: {', '.join(sorted(clashes))}"

    if errors:
        return templates.TemplateResponse(
            "purchase/receive.html",
            {"request": request, "user": user, "order": order, "errors": errors, "submitted": submitted},
            status_code=400,
        )

    # Everything validated — now write stock, serials, and status together.
    for line in order.items:
        item = db.query(Item).get(line.item_id)
        item.current_stock = (item.current_stock or 0) + line.quantity
        db.add(StockTransaction(
            item_id=item.id, transaction_type="IN", quantity=line.quantity,
            reference_type="purchase_order", reference_id=order.id,
            notes=f"Received against {order.order_no}",
        ))
        for serial in serials_by_line.get(line.id, []):
            db.add(ItemSerial(
                item_id=item.id, serial_number=serial,
                purchase_order_id=order.id, purchase_order_item_id=line.id,
            ))

    order.status = "received"
    log_action(
        db, user, "purchase_order", order.id, order.order_no, "receive",
        summary=f"Received — {len(order.items)} line(s), stock and serials updated",
    )
    db.commit()
    return RedirectResponse(url=f"/purchase/{order_id}?success=Order received and stock updated", status_code=303)


def get_purchase_order_delete_blockers(order: PurchaseOrder) -> list[str]:
    """Reasons a purchase order can't be hard-deleted — same style as
    items.py's get_item_delete_blockers. 'received' has real history
    depending on it: ItemSerial rows carry an actual FK to this order
    (purchase_order_id/purchase_order_item_id), and StockTransaction rows
    reference it too (reference_id — not a DB-level FK, but still meant to
    stay resolvable as an audit trail) — deleting it would either violate
    the serial FK outright or silently orphan the stock history. 'ordered'
    isn't a technical/FK problem (nothing writes stock until receipt), but
    it represents a real, admin-approved and confirmed document that may
    already be with the vendor — Cancel is the correct way to retire one of
    those without losing the record of what was approved. 'draft',
    'pending_approval', and 'cancelled' never touched any of that, so those
    stay freely deletable. 'pending_confirmation' is deliberately NOT
    listed as freely deletable either — it's already been through admin
    approval, so it falls into the same "cancel instead" bucket as 'ordered'
    below rather than being treated like a plain draft."""
    blockers = []
    if order.status == "received":
        blockers.append("it has been received — stock levels and serial numbers were recorded against it")
    elif order.status not in ("draft", "pending_approval", "cancelled"):
        blockers.append(f"it's been approved and is '{order.status}' — cancel it instead to keep a record of what was ordered")
    return blockers


@router.post("/{order_id}/delete")
def delete_purchase_order(order_id: int, db: Session = Depends(get_db), user: User = Depends(require_module_permission("purchase", "delete"))):
    order = db.query(PurchaseOrder).options(joinedload(PurchaseOrder.vendor)).get(order_id)
    if not order:
        return RedirectResponse(url="/purchase?error=Purchase order not found", status_code=303)

    blockers = get_purchase_order_delete_blockers(order)
    if blockers:
        msg = f"Can't delete {order.order_no} — {'; '.join(blockers)}."
        return RedirectResponse(url=f"/purchase/{order_id}?error={quote(msg)}", status_code=303)

    try:
        log_action(
            db, user, "purchase_order", order.id, order.order_no, "delete",
            summary=f"Deleted '{order.order_no}' ({order.vendor.name if order.vendor else order.vendor_id})",
        )
        db.delete(order)
        db.commit()
    except IntegrityError:
        # Belt-and-braces: catch any reference the checks above didn't
        # anticipate rather than surfacing a raw 500 to the user.
        db.rollback()
        msg = f"Can't delete {order.order_no} — it's still referenced elsewhere."
        return RedirectResponse(url=f"/purchase/{order_id}?error={quote(msg)}", status_code=303)

    return RedirectResponse(url="/purchase?success=Purchase order deleted", status_code=303)
